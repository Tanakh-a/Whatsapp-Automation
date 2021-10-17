import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


class successfullySent:

    # def write_sent(self):
    #     header = ['Status']
    #     workbook_name = "E:/Project/Excel/MsgSent.xlsx"
    #     wb = Workbook()
    #     page = wb.active
    #     page.title = 'Result'
    #     page.append(header)  # write header to the first line
    #
    #     result = [['SENT']]
    #
    #     for res in result:
    #         page.append(res)
    #     wb.save(filename=workbook_name)

    def update_ws(self):
        wb_name = "E:/Project/Excel/MsgSent.xlsx"
        ws = load_workbook(wb_name)
        wb = Workbook()

        page = wb.active
        result = ['SENT']


        row = page.max_row
        for val in result:
            page.cell(row=row + 1, column=1, value=val)
            page.append(result)
            # page.cell(c=c + col, row=1, value=result)
        # for res in result:
        #     page.append(res)
        wb.save(filename=wb_name)

        # j = 1
        # i = 1
        # while i > 0:
        #     if page.cell(row=j, column=1).value is None:
        #         page.append(['TEST'])
        #         wb.save("MsgSent.xlsx")
        #         break
        #     else:
        #         j = j + 1



#ins.write_sent()
# ins = successfullySent()
# ins.update_ws()
