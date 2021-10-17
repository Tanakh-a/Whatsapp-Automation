from openpyxl import Workbook, load_workbook


class dataRead():

    def contact(self):
        wb = load_workbook(filename='E:/Project/Excel/excel.xlsx')
        sh = wb.active
        name = sh.cell(3, 1).value
        print(name)
        return name
