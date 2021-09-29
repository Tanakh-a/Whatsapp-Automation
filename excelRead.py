# Module file: exelRead.py
from  openpyxl import Workbook, load_workbook
wb = load_workbook(filename='excel.xlsx')
sh = wb.active
#print(sh['A2'].value)

contact_no = (sh.cell(3,2).value)
print(sh.cell(3,2).value)

