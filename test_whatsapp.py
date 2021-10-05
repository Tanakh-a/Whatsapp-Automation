import time
import unittest
import HtmlTestRunner
import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys


class WhatsApp(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.driver = webdriver.Chrome()
        cls.driver.implicitly_wait(10)
        cls.driver.maximize_window()

    def test_readData(self):
        # import openpyxl
        path = '/home/tanakh/PycharmProjects/whatsapp-testing/PySelenium/excel.xlsx'
        wb = openpyxl.load_workbook(path)
        sh = wb['number']
        print(sh['B3'].value)
        return str(sh['B3'].value)

    def test_send_messege(self, number, message):
        # driver = webdriver.Chrome()
        self.driver.get("https://web.whatsapp.com/")
        time.sleep(15)
        search_box = self.driver.find_element_by_xpath('//*[@id="side"]/div[1]/div/label')
        search_box.send_keys()
        time.sleep(2)
        self.driver.find_element_by_xpath('//*[@id="pane-side"]/div[1]/div/div/div[9]/div/div/div[2]').click()

        time.sleep(2)
        message_box = self.driver.find_element_by_xpath(
            '//*[@id="main"]/footer/div[1]/div/div/div[2]/div[1]/div/div[2]')
        message_box.send_keys("Sent from bot.Don't reply...")
        message_box.send_keys(Keys.ENTER)
        time.sleep(10)

        number = test_readData(self)

        message = "Sent from bot.Don't reply..."
        print("test_num: " + number)
        instance1 = WhatsApp(unittest.TestCase)
        instance1.test_send_messege(number, message)

    def test_result(self):
        # from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # test_sent = [["sent"]]
        ws.cell(row=1, column=1).value = 'SENT_unsent'
        # for res in test_sent:
        #     ws.write(test_sent)
        wb.save("sent.xlsx")

    def test_logout(self):
        menu = self.driver.find_element_by_xpath('//*[@id="side"]/header/div[2]/div/span/div[3]/div/span').click


        time.sleep(8)
        logout = self.driver.find_element_by_xpath('//*[@id="side"]/header/div[2]/div/span/div[3]/span/div[1]/ul/li[5]/div[1]').click


    @classmethod
    def tearDownClass(cls):
        cls.driver.close()
        # self.driver.quit()
        print("Test Completed")


if __name__ == '__main__':
    unittest.main(
        testRunner=HtmlTestRunner.HTMLTestRunner(output='/home/tanakh/PycharmProjects/whatsapp-testing/report'))
